from sqlalchemy.orm import Session
from sqlalchemy import func
import openpyxl
from io import BytesIO
from app.models.organization import Organization
from app.schemas.organization import OrganizationCreate, OrganizationUpdate
from typing import List, Optional


def generate_organization_id(db: Session) -> str:
    max_id = db.query(func.max(Organization.id)).scalar() or 0
    return f"ORG{max_id + 1:03d}"

def create_organization(db: Session, payload: OrganizationCreate) -> Organization:
    org_id = generate_organization_id(db)
    new_org = Organization(**payload.model_dump(), organization_id=org_id)
    db.add(new_org)
    db.commit()
    db.refresh(new_org)
    return new_org


def get_all_organizations(
    db: Session, 
    status: Optional[str] = None, 
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> List[Organization]:
    # Part 5: Filter by is_active = 1
    query = db.query(Organization).filter(Organization.is_active == 1)
    
    if status:
        query = query.filter(Organization.status == status.lower())
    
    if search:
        search_term = f"%{search.strip().lower()}%"
        query = query.filter(func.lower(Organization.organization_name).like(search_term))
        
    if start_date:
        query = query.filter(Organization.contract_end_date >= start_date)
        
    if end_date:
        query = query.filter(Organization.contract_end_date <= end_date)
        
    return query.order_by(Organization.organization_name.asc()).all()


def export_organizations_to_excel(organizations: List[Organization]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organizations"

    # Header
    headers = ["DB ID", "Organization ID", "Organization Name", "Contact Number", "Country Code", "Address", "Commission %", "Status", "Contract Signed Date", "Contract End Date", "Created At", "Updated At"]
    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    for org in organizations:
        ws.append([
            org.id,
            org.organization_id,
            org.organization_name,
            org.contact_number or "",
            org.country_code or "",
            org.address or "",
            float(org.commission_percentage) if org.commission_percentage else 0.0,
            org.status,
            org.contract_signed_date.strftime("%Y-%m-%d") if org.contract_signed_date else "",
            org.contract_end_date.strftime("%Y-%m-%d") if org.contract_end_date else "",
            org.created_at.strftime("%Y-%m-%d %H:%M:%S") if org.created_at else "",
            org.updated_at.strftime("%Y-%m-%d %H:%M:%S") if org.updated_at else ""
        ])

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_organization_by_id(db: Session, org_id: int) -> Optional[Organization]:
    return db.query(Organization).filter(Organization.id == org_id).first()


def update_organization(db: Session, org_id: int, payload: OrganizationUpdate) -> Optional[Organization]:
    db_org = get_organization_by_id(db, org_id)
    if not db_org:
        return None
    
    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_org, key, value)
    
    db.commit()
    db.refresh(db_org)
    return db_org


def delete_organization(db: Session, org_id: int) -> bool:
    db_org = get_organization_by_id(db, org_id)
    if not db_org:
        return False
    db.delete(db_org)
    db.commit()
    return True


def check_organization_exists(db: Session, name: str) -> bool:
    """Checks if organization exists with case-insensitive name."""
    clean_name = name.strip().lower()
    print(f"DEBUG: Checking organization duplicate: '{clean_name}'")
    
    exists = db.query(Organization).filter(
        func.lower(func.trim(Organization.organization_name)) == clean_name
    ).first()
    
    print(f"DEBUG: Duplicate check result: {'EXISTS' if exists else 'NOT FOUND'}")
    return exists is not None
