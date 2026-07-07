from typing import List, Optional
from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from app.models.employee import Employee
from app.schemas.employee import EmployeeCreateRequest, EmployeeUpdateRequest

# ─────────────────────────────────────────────────────────────
# INTERNAL HELPER
# ─────────────────────────────────────────────────────────────

def generate_employee_id(db: Session) -> str:
    """
    Auto-generates the next employee ID in the format RM0001.
    """
    # Find the maximum 'id' in the table
    max_id = db.query(func.max(Employee.id)).scalar() or 0
    next_number = max_id + 1
    return f"RM{next_number:04d}"

def compute_employee_completion(employee: Employee):
    """
    Computes and sets the completion_percentage and profile_status for an employee.
    Calculates HR completion and Admin completion separately by counting individual fields.
    """
    # Helper to check if a value is non-empty
    def is_filled(val):
        return val is not None and str(val).strip() != ""

    # 1. HR Details Fields (38 fields, plus 1 if Inactive)
    hr_fields = [
        employee.first_name,
        employee.last_name,
        employee.gender,
        employee.blood_group,
        employee.date_of_birth,
        employee.email,
        employee.contact_number,
        employee.contact_number_office,
        employee.emergency_contact_number,
        employee.medical_condition,
        employee.current_address,
        employee.present_address_proof_url,
        employee.permanent_address,
        employee.permanent_address_proof_url,
        employee.aadhar_number,
        employee.aadhar_url,
        employee.pan_number,
        employee.pan_url,
        employee.marksheet_10th_url,
        employee.marksheet_12th_url,
        employee.marksheet_graduation_url,
        employee.photo_url,
        employee.package,
        employee.date_of_joining,
        employee.status,
        employee.last_company_name,
        employee.resume_url,
        employee.salary_slips_url,
        employee.offer_letter_url,
        employee.designation,
        employee.assigned_business_unit,
        employee.reporting_to,
        employee.work_mode,
        employee.ctc,
        employee.compliance,
        employee.bank_name,
        employee.bank_account_number,
        employee.bank_ifsc_code
    ]
    if employee.status == 'Inactive':
        hr_fields.append(employee.last_working_date)

    total_hr = len(hr_fields)
    filled_hr = sum(1 for val in hr_fields if is_filled(val))
    hr_pct = int((filled_hr / total_hr) * 100) if total_hr > 0 else 0
    
    employee.completion_percentage_hr = hr_pct
    
    if hr_pct == 100:
        employee.profile_status_hr = "Completed"
    elif hr_pct > 0:
        employee.profile_status_hr = "In Progress"
    else:
        employee.profile_status_hr = "Draft"

    # 2. ADMIN Details Fields (6 fields)
    admin_fields = [
        employee.system_assigned, 
        employee.sim_card_assigned, 
        employee.email_id_configured, 
        employee.linkedin_configured, 
        employee.google_sheet_configured, 
        employee.whatsapp_business_configured
    ]
    total_admin = len(admin_fields)
    filled_admin = sum(1 for val in admin_fields if is_filled(val))
    admin_pct = int((filled_admin / total_admin) * 100) if total_admin > 0 else 0
    
    employee.completion_percentage_admin = admin_pct
    
    if admin_pct == 100:
        employee.profile_status_admin = "Completed"
    elif admin_pct > 0:
        employee.profile_status_admin = "In Progress"
    else:
        employee.profile_status_admin = "Draft"

    # 3. Overall Completion Percentage (weighted average: 90% HR + 10% Admin)
    pct = int(hr_pct * 0.9 + admin_pct * 0.1)
    employee.completion_percentage = pct
    
    if pct == 100:
        employee.profile_status = "Completed"
    elif pct > 0:
        employee.profile_status = "In Progress"
    else:
        employee.profile_status = "Draft"

    # Generate password if both sections are 100% complete
    if employee.completion_percentage_hr == 100 and employee.completion_percentage_admin == 100:
        import re
        # Extract digits from employee_id (e.g. RM0001 -> 0001)
        digits = "".join(re.findall(r"\d+", employee.employee_id or ""))
        first_char = employee.first_name[0] if employee.first_name else ""
        last_name = employee.last_name or ""
        # Format: first_char + last_name + "@" + digits
        employee.employee_password = f"{first_char}{last_name}@{digits}"
    else:
        employee.employee_password = None
        
# ─────────────────────────────────────────────────────────────
# CREATE
# ─────────────────────────────────────────────────────────────

def create_employee(db: Session, payload: EmployeeCreateRequest) -> Employee:
    """
    Creates a new employee record.
    """
    emp_code = generate_employee_id(db)
    
    new_employee = Employee(
        employee_id=emp_code,
        date=date.today(),
        first_name=payload.first_name,
        last_name=payload.last_name,
        blood_group=payload.blood_group,
        gender=payload.gender,
        country_code=payload.country_code,
        contact_number=payload.contact_number,
        email=payload.email,
        permanent_address=payload.permanent_address,
        current_address=payload.current_address,
        designation=payload.designation,
        date_of_joining=payload.date_of_joining,
        package=payload.package,
        status=payload.status,
        last_working_date=payload.last_working_date,
        
        # New fields
        date_of_birth=payload.date_of_birth,
        countrycode_office_contact=payload.countrycode_office_contact,
        contact_number_office=payload.contact_number_office,
        countrycode_emergency_contact=payload.countrycode_emergency_contact,
        emergency_contact_number=payload.emergency_contact_number,
        aadhar_number=payload.aadhar_number,
        aadhar_url=payload.aadhar_url,
        pan_number=payload.pan_number,
        pan_url=payload.pan_url,
        marksheet_10th_url=payload.marksheet_10th_url,
        marksheet_12th_url=payload.marksheet_12th_url,
        marksheet_graduation_url=payload.marksheet_graduation_url,
        present_address_proof_url=payload.present_address_proof_url,
        permanent_address_proof_url=payload.permanent_address_proof_url,
        photo_url=payload.photo_url,
        medical_condition=payload.medical_condition,
        resume_url=payload.resume_url,
        salary_slips_url=payload.salary_slips_url,
        offer_letter_url=payload.offer_letter_url,
        last_company_name=payload.last_company_name,
        
        # Bank Details
        bank_name=payload.bank_name,
        bank_account_number=payload.bank_account_number,
        bank_ifsc_code=payload.bank_ifsc_code,

        # Reporting & Compliance Details
        assigned_business_unit=payload.assigned_business_unit,
        reporting_to=payload.reporting_to,
        work_mode=payload.work_mode,
        ctc=payload.ctc,
        compliance=payload.compliance,

        # Asset & System Configuration Details
        system_assigned=payload.system_assigned,
        sim_card_assigned=payload.sim_card_assigned,
        email_id_configured=payload.email_id_configured,
        linkedin_configured=payload.linkedin_configured,
        google_sheet_configured=payload.google_sheet_configured,
        whatsapp_business_configured=payload.whatsapp_business_configured,
        employee_password=payload.employee_password
    )
    
    compute_employee_completion(new_employee)
    
    db.add(new_employee)
    db.commit()
    db.refresh(new_employee)
    return new_employee

# ─────────────────────────────────────────────────────────────
# READ (LIST ALL)
# ─────────────────────────────────────────────────────────────

def get_all_employees(
    db: Session, 
    search: Optional[str] = None, 
    status: Optional[str] = None,
    designation: Optional[str] = None,
    min_package: Optional[float] = None,
    max_package: Optional[float] = None,
    blood_group: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc"
) -> List[Employee]:
    """
    Fetches all employees with optional search and filters.
    """
    query = db.query(Employee)

    if search:
        search_term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Employee.employee_id.ilike(search_term),
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.designation.ilike(search_term)
            )
        )

    if status and status.upper() != "ALL":
        query = query.filter(Employee.status == status)
        
    if designation:
        query = query.filter(Employee.designation == designation)
        
    if min_package is not None:
        query = query.filter(Employee.package >= min_package)
        
    if max_package is not None:
        query = query.filter(Employee.package <= max_package)
        
    if blood_group:
        query = query.filter(Employee.blood_group == blood_group)

    from app.utils.sorting import apply_sorting
    query = apply_sorting(query, Employee, sort_by, sort_order, Employee.id)
    return query.all()

import openpyxl
from io import BytesIO

def export_employees_to_excel(employees: List[Employee]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Header containing all db fields
    headers = [
        "ID", "Employee ID", "First Name", "Last Name",
        "Blood Group", "Gender", "Country Code", "Contact Number (Personal)", "Email",
        "Permanent Address", "Current Address (Present)", "Designation", "Date of Joining", "Creation Date",
        "Package (LPA)", "Status", "Last Working Date", "Date of Birth", "Office Country Code", "Contact Number (Office)",
        "Emergency Country Code", "Emergency Contact", "Aadhar Number", "Aadhar URL", "PAN Number", "PAN URL",
        "10th Marksheet", "12th Marksheet", "Graduation Marksheet", "Present Address Proof",
        "Permanent Address Proof", "Photo", "Medical Condition",
        "Resume", "Salary Slips (Last 3)", "Last Offer Letter", "Last Company Name",
        "Assigned Business Unit", "Reporting To", "Work Mode", "CTC", "Compliance",
        "System Assigned", "SIM Card Assigned", "Email ID Configured", "LinkedIn Configured", "Google Sheet Configured", "Whatsapp Business Configuration",
        "HR Profile Status", "HR Completion %", "Admin Profile Status", "Admin Completion %", "Employee Password"
    ]
    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    for emp in employees:
        ws.append([
            emp.id,
            emp.employee_id,
            emp.first_name,
            emp.last_name,
            emp.blood_group or "",
            emp.gender or "",
            emp.country_code or "",
            emp.contact_number or "",
            emp.email or "",
            emp.permanent_address or "",
            emp.current_address or "",
            emp.designation,
            emp.date_of_joining.strftime("%Y-%m-%d") if emp.date_of_joining else "",
            emp.date.strftime("%Y-%m-%d") if emp.date else "",
            float(emp.package) if emp.package else 0.0,
            emp.status,
            emp.last_working_date.strftime("%Y-%m-%d") if emp.last_working_date else "",
            emp.date_of_birth.strftime("%Y-%m-%d") if emp.date_of_birth else "",
            emp.countrycode_office_contact or "",
            emp.contact_number_office or "",
            emp.countrycode_emergency_contact or "",
            emp.emergency_contact_number or "",
            emp.aadhar_number or "",
            emp.aadhar_url or "",
            emp.pan_number or "",
            emp.pan_url or "",
            emp.marksheet_10th_url or "",
            emp.marksheet_12th_url or "",
            emp.marksheet_graduation_url or "",
            emp.present_address_proof_url or "",
            emp.permanent_address_proof_url or "",
            emp.photo_url or "",
            emp.medical_condition or "",
            emp.resume_url or "",
            emp.salary_slips_url or "",
            emp.offer_letter_url or "",
            emp.last_company_name or "",
            emp.assigned_business_unit or "",
            emp.reporting_to or "",
            emp.work_mode or "",
            float(emp.ctc) if emp.ctc is not None else 0.0,
            emp.compliance or "",
            emp.system_assigned or "",
            emp.sim_card_assigned or "",
            emp.email_id_configured or "",
            emp.linkedin_configured or "",
            emp.google_sheet_configured or "",
            emp.whatsapp_business_configured or "",
            emp.profile_status_hr or "Draft",
            emp.completion_percentage_hr or 0,
            emp.profile_status_admin or "Draft",
            emp.completion_percentage_admin or 0,
            emp.employee_password or ""
        ])

    # Adjust column widths
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ─────────────────────────────────────────────────────────────
# READ (GET ONE)
# ─────────────────────────────────────────────────────────────

def get_employee_by_id(db: Session, db_id: int) -> Optional[Employee]:
    """
    Fetches a single employee by their database primary key (id).
    """
    return db.query(Employee).filter(Employee.id == db_id).first()

# ─────────────────────────────────────────────────────────────
# UPDATE
# ─────────────────────────────────────────────────────────────

def update_employee(
    db: Session, 
    db_id: int, 
    payload: EmployeeUpdateRequest
) -> Optional[Employee]:
    """
    Updates an existing employee. Only updates fields that were provided.
    """
    employee = db.query(Employee).filter(Employee.id == db_id).first()
    
    if not employee:
        return None

    # payload.model_dump(exclude_unset=True) returns a dictionary of only 
    # the fields the frontend actually sent.
    update_data = payload.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(employee, key, value)
        
    compute_employee_completion(employee)

    db.commit()
    db.refresh(employee)
    return employee

# ─────────────────────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────────────────────

def delete_employee(db: Session, db_id: int) -> bool:
    """
    Deletes an employee from the database.
    """
    employee = db.query(Employee).filter(Employee.id == db_id).first()
    
    if not employee:
        return False
        
    db.delete(employee)
    db.commit()
    return True
