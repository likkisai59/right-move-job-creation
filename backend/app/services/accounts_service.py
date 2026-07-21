from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.account import Account
from app.models.employee import Employee
from app.models.leave import Leave
from app.models.payroll_config import PayrollConfig
from app.models.invoice import Invoice
from app.models.job_candidate import JobCandidateMapping
from app.models.candidate import Candidate
from app.models.job_requirement import Job
from app.models.organization import Organization
from app.schemas.account import AccountCreate, AccountUpdate
from app.schemas.invoice import InvoiceUpdate
from typing import List, Tuple, Optional
from datetime import date

# ── PAYROLL CONFIGURATION SERVICES ───────────────────────────

def get_payroll_config(db: Session) -> PayrollConfig:
    """
    Retrieve the payroll configuration. Creates a default row if none exists.
    """
    config = db.query(PayrollConfig).first()
    if not config:
        config = PayrollConfig(pf_percentage=12.0, tds_percentage=10.0)
        db.add(config)
        db.commit()
        db.refresh(config)
    return config

def update_payroll_config(db: Session, pf_percentage: float, tds_percentage: float) -> PayrollConfig:
    """
    Update the PF and TDS configuration percentages.
    """
    config = get_payroll_config(db)
    config.pf_percentage = pf_percentage
    config.tds_percentage = tds_percentage
    db.commit()
    db.refresh(config)
    return config

# ── CALCULATION UTILITIES ─────────────────────────────────────

def calculate_unpaid_leave_amount_and_net_payable(basic_pay: int, unpaid_leaves: float) -> Tuple[int, int]:
    """
    Calculate Leave Amount and Net Payable Salary.
    Leave Amount = (basic salary // 30) * unpaid leaves
    Net Payable Salary = basic salary - Leave Amount
    """
    unpaid_leave_amount = int((basic_pay // 30) * unpaid_leaves)
    net_payable_salary = basic_pay - unpaid_leave_amount
    return unpaid_leave_amount, net_payable_salary

def get_current_salary_cycle_range(today: date, date_of_joining: Optional[date]) -> Tuple[date, date]:
    """
    Returns (start_date, end_date) for the current salary cycle.
    - If today is <= 25:
      - The cycle ends on the 25th of the current month.
      - The cycle starts on the 26th of the previous month.
    - If today is >= 26:
      - The cycle ends on the 25th of the next month.
      - The cycle starts on the 26th of the current month.
    If the employee joined after the standard start_date, the cycle starts on the date of joining.
    """
    from datetime import timedelta
    if today.day <= 25:
        end_date = date(today.year, today.month, 25)
        first_of_current = date(today.year, today.month, 1)
        prev_month_end = first_of_current - timedelta(days=1)
        start_date = date(prev_month_end.year, prev_month_end.month, 26)
    else:
        if today.month == 12:
            end_date = date(today.year + 1, 1, 25)
        else:
            end_date = date(today.year, today.month + 1, 25)
        start_date = date(today.year, today.month, 26)
        
    if date_of_joining and date_of_joining > start_date:
        start_date = date_of_joining
        
    return start_date, end_date

def get_recruiter_incentives_sum(db: Session, employee: Employee, start_date: Optional[date] = None, end_date: Optional[date] = None) -> int:
    """
    Sum incentives for approved/joined candidates recruited by this employee.
    Matches Candidate.recruiter_name case-insensitively with Employee full name.
    Optionally filters by Candidate approval date within cycle range.
    """
    emp_name = f"{employee.first_name} {employee.last_name}".strip()
    
    query = db.query(JobCandidateMapping).join(Candidate).filter(
        func.lower(Candidate.recruiter_name) == func.lower(emp_name),
        JobCandidateMapping.status.in_(["Candidate Approved", "Joined"])
    )
    
    if start_date and end_date:
        query = query.filter(
            JobCandidateMapping.approval_date >= start_date,
            JobCandidateMapping.approval_date <= end_date
        )
        
    mappings = query.all()
    
    total_incentives_sum = 0
    for m in mappings:
        if m.incentive:
            try:
                clean_inc = "".join(c for c in m.incentive if c.isdigit())
                if clean_inc:
                    total_incentives_sum += int(clean_inc)
            except ValueError:
                pass
    return total_incentives_sum

# ── ACCOUNTS MODULE CRUD ──────────────────────────────────────

def list_accounts(db: Session) -> List[Account]:
    # Query all active employees
    employees = db.query(Employee).filter(Employee.status == "Active").all()
    config = get_payroll_config(db)
    
    result = []
    for emp in employees:
        account = db.query(Account).filter(Account.employee_id == emp.id).first()
        
        # Determine current salary cycle dates
        joining_date = emp.date or emp.date_of_joining
        start_date, end_date = get_current_salary_cycle_range(date.today(), joining_date)
        
        # Calculate sum of approved unpaid leaves within current cycle
        unpaid_leaves_sum = db.query(func.sum(Leave.total_leaves)).filter(
            Leave.employee_id == emp.id,
            Leave.status == "Approved",
            Leave.leave_type.ilike("%unpaid%"),
            Leave.start_date >= start_date,
            Leave.start_date <= end_date
        ).scalar() or 0.0
        
        # Calculate recruiter incentives within current cycle
        total_incentives_sum = get_recruiter_incentives_sum(db, emp, start_date, end_date)
        
        # CTC Offered from employees table
        ctc_val = int(emp.ctc) if emp.ctc is not None else 0
        
        if account:
            calculated_unpaid_leave_amount, calculated_net_payable = calculate_unpaid_leave_amount_and_net_payable(account.basic_pay, unpaid_leaves_sum)
            calculated_total_net_payable = calculated_net_payable + total_incentives_sum
            calculated_gross_salary = calculated_total_net_payable + (account.additional_incentive or 0) + (account.client_incentive or 0) - account.deduction_amount
            
            # Dynamic PF and TDS calculations based on compliance type
            calculated_pf = 0.0
            calculated_tds = 0.0
            compliance_upper = emp.compliance.strip().upper() if emp.compliance else ""
            if compliance_upper == "PF":
                calculated_pf = float(((calculated_net_payable or 0) + (account.hra or 0)) * (config.pf_percentage / 100.0))
            elif compliance_upper == "TDS":
                calculated_tds = float(((calculated_net_payable or 0) + (account.hra or 0)) * (config.tds_percentage / 100.0))
                
            # Net Salary Payout = Calculated Basic Pay + HRA - PF - TDS + Additional Incentive + Client Incentive - Incentive Deducted + Candidate Incentives - Loan Deducted
            calculated_total_gross = int(
                calculated_net_payable + 
                (account.hra or 0) - 
                calculated_pf - 
                calculated_tds + 
                (account.additional_incentive or 0) + 
                (account.client_incentive or 0) - 
                (account.incentive_deducted or 0) + 
                total_incentives_sum - 
                (account.loan_deducted or 0)
            )
            
            # Sync to DB if stored value is different
            calculated_baseline_status = (50 if (account.basic_pay or 0) > 0 else 0) + (50 if (account.hra or 0) > 0 else 0)
            
            if (account.unpaid_leaves != unpaid_leaves_sum or 
                account.unpaid_leave_amount != calculated_unpaid_leave_amount or 
                account.net_payable_salary != calculated_net_payable or
                account.incentives != total_incentives_sum or
                getattr(account, 'candidate_incentives', 0) != total_incentives_sum or
                account.ctc_offered != ctc_val or
                account.total_net_payable_salary != calculated_total_net_payable or
                account.gross_salary != calculated_gross_salary or
                account.pf != calculated_pf or
                account.tds != calculated_tds or
                account.net_salary_pay != calculated_total_gross or
                account.calculated_basic_pay != calculated_net_payable or
                getattr(account, 'baseline_status', 0) != calculated_baseline_status):
                
                account.unpaid_leaves = unpaid_leaves_sum
                account.unpaid_leave_amount = calculated_unpaid_leave_amount
                account.net_payable_salary = calculated_net_payable
                account.incentives = total_incentives_sum
                if hasattr(account, 'candidate_incentives'):
                    account.candidate_incentives = total_incentives_sum
                account.ctc_offered = ctc_val
                account.total_net_payable_salary = calculated_total_net_payable
                account.gross_salary = calculated_gross_salary
                account.pf = calculated_pf
                account.tds = calculated_tds
                account.net_salary_pay = calculated_total_gross
                account.calculated_basic_pay = calculated_net_payable
                if hasattr(account, 'baseline_status'):
                    account.baseline_status = calculated_baseline_status
                db.commit()
                db.refresh(account)
            result.append(account)
        else:
            # Construct a default transient Account object linked to this employee
            calculated_total_net_payable = 0 + total_incentives_sum
            calculated_gross_salary = calculated_total_net_payable - 0
            
            calculated_pf = 0.0
            calculated_tds = 0.0
            compliance_upper = emp.compliance.strip().upper() if emp.compliance else ""
            if compliance_upper == "PF":
                calculated_pf = float((0 + 0) * (config.pf_percentage / 100.0))
            elif compliance_upper == "TDS":
                calculated_tds = float((0 + 0) * (config.tds_percentage / 100.0))
                
            # Net Salary Payout = Calculated Basic Pay + HRA - PF - TDS + Additional Incentive + Client Incentive - Incentive Deducted + Candidate Incentives - Loan Deducted
            calculated_total_gross = int(
                0 + 
                0 - 
                calculated_pf - 
                calculated_tds + 
                0 + 
                0 - 
                0 + 
                total_incentives_sum - 
                0
            )
            
            default_account = Account(
                id=None,
                employee_id=emp.id,
                basic_pay=0,
                hra=0,
                loan_amount=0,
                client_incentive=0,
                deduction_amount=0,
                unpaid_leaves=unpaid_leaves_sum,
                unpaid_leave_amount=0,
                net_payable_salary=0,
                ctc_offered=ctc_val,
                incentives=total_incentives_sum,
                candidate_incentives=total_incentives_sum,
                client_total=0,
                total_net_payable_salary=calculated_total_net_payable,
                gross_salary=calculated_gross_salary,
                pf=calculated_pf,
                tds=calculated_tds,
                additional_incentive=0,
                incentive_deducted=0,
                loan_deducted=0,
                net_salary_pay=calculated_total_gross,
                calculated_basic_pay=0,
                baseline_status=0
            )
            default_account.employee = emp
            result.append(default_account)
            
    return result

def create_account(db: Session, payload: AccountCreate) -> Account:
    employee = db.query(Employee).filter(Employee.id == payload.employee_id).first()
    if not employee:
        raise ValueError("Employee not found")

    existing_account = db.query(Account).filter(Account.employee_id == payload.employee_id).first()
    if existing_account:
        raise ValueError("Account details already exist for this employee")

    new_account = Account(
        **payload.model_dump()
    )
    config = get_payroll_config(db)
    
    new_account.unpaid_leave_amount, new_account.net_payable_salary = calculate_unpaid_leave_amount_and_net_payable(
        new_account.basic_pay, new_account.unpaid_leaves
    )
    new_account.calculated_basic_pay = new_account.net_payable_salary
    
    new_account.ctc_offered = int(employee.ctc) if employee.ctc is not None else 0
    new_account.incentives = get_recruiter_incentives_sum(db, employee)
    
    new_account.total_net_payable_salary = new_account.net_payable_salary + new_account.incentives
    new_account.gross_salary = int(new_account.total_net_payable_salary + (new_account.additional_incentive or 0) + (new_account.client_incentive or 0) - new_account.deduction_amount)
    
    # Calculate PF/TDS and Total Gross
    compliance_upper = employee.compliance.strip().upper() if employee.compliance else ""
    if compliance_upper == "PF":
        new_account.pf = float(((new_account.calculated_basic_pay or 0) + (new_account.hra or 0)) * (config.pf_percentage / 100.0))
        new_account.tds = 0.0
    elif compliance_upper == "TDS":
        new_account.pf = 0.0
        new_account.tds = float(((new_account.calculated_basic_pay or 0) + (new_account.hra or 0)) * (config.tds_percentage / 100.0))
    else:
        new_account.pf = 0.0
        new_account.tds = 0.0
        
    new_account.net_salary_pay = int((new_account.gross_salary or 0) - (new_account.pf or 0) - (new_account.tds or 0) - (new_account.loan_deducted or 0) - (new_account.incentive_deducted or 0))
    new_account.baseline_status = (50 if (new_account.basic_pay or 0) > 0 else 0) + (50 if (new_account.hra or 0) > 0 else 0)
    
    db.add(new_account)
    db.commit()
    db.refresh(new_account)
    return new_account

def get_account_by_employee(db: Session, employee_id: int) -> Optional[Account]:
    return db.query(Account).filter(Account.employee_id == employee_id).first()

def update_account(db: Session, employee_id: int, payload: AccountUpdate) -> Account:
    account = db.query(Account).filter(Account.employee_id == employee_id).first()
    if not account:
        raise ValueError("Account details not found for this employee")

    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(account, key, value)

    account.unpaid_leave_amount, account.net_payable_salary = calculate_unpaid_leave_amount_and_net_payable(
        account.basic_pay, account.unpaid_leaves
    )
    account.calculated_basic_pay = account.net_payable_salary

    employee = db.query(Employee).filter(Employee.id == account.employee_id).first()
    if employee:
        account.ctc_offered = int(employee.ctc) if employee.ctc is not None else 0
        account.incentives = get_recruiter_incentives_sum(db, employee)

    account.total_net_payable_salary = account.net_payable_salary + account.incentives
    account.gross_salary = int(account.total_net_payable_salary + (account.additional_incentive or 0) + (account.client_incentive or 0) - account.deduction_amount)

    # Recalculate PF/TDS and Total Gross
    config = get_payroll_config(db)
    if employee:
        compliance_upper = employee.compliance.strip().upper() if employee.compliance else ""
        if compliance_upper == "PF":
            account.pf = float(((account.calculated_basic_pay or 0) + (account.hra or 0)) * (config.pf_percentage / 100.0))
            account.tds = 0.0
        elif compliance_upper == "TDS":
            account.pf = 0.0
            account.tds = float(((account.calculated_basic_pay or 0) + (account.hra or 0)) * (config.tds_percentage / 100.0))
        else:
            account.pf = 0.0
            account.tds = 0.0
            
    account.net_salary_pay = int((account.gross_salary or 0) - (account.pf or 0) - (account.tds or 0) - (account.loan_deducted or 0) - (account.incentive_deducted or 0))
    account.baseline_status = (50 if (account.basic_pay or 0) > 0 else 0) + (50 if (account.hra or 0) > 0 else 0)

    db.commit()
    db.refresh(account)
    return account

# ── SECOND TAB: PLACEMENT RECORDS ───────────────────────────────
def list_placements(db: Session) -> List[dict]:
    """
    List joined placements from the Job Candidate mapping module.
    """
    mappings = db.query(JobCandidateMapping).filter(
        JobCandidateMapping.status == "Joined"
    ).all()
    
    placements = []
    for m in mappings:
        candidate = m.candidate
        job = m.job
        
        # Get organization details through Job
        org = None
        if job.organization_id:
            org = db.query(Organization).filter(Organization.id == job.organization_id).first()
            
        # Look up employee details for the recruiter
        rec_emp_id = "—"
        rec_emp_name = candidate.recruiter_name or "—"
        if candidate.recruiter_name:
            search_name = candidate.recruiter_name.strip().replace(" ", "").lower()
            employees_list = db.query(Employee).all()
            for emp in employees_list:
                emp_full = f"{emp.first_name or ''}{emp.last_name or ''}".replace(" ", "").lower()
                if emp_full == search_name:
                    rec_emp_id = emp.employee_id
                    rec_emp_name = f"{emp.first_name or ''} {emp.last_name or ''}".strip()
                    break

        placements.append({
            "approval_date": m.approval_date,
            "candidate_code": candidate.candidate_code,
            "candidate_name": f"{candidate.first_name} {candidate.last_name}".strip(),
            "organization_id": org.organization_id if org else None,
            "organization_name": org.organization_name if org else job.company_name,
            "location": org.location if org else None,
            "job_designation": job.requirements[0].job_title if job.requirements else "—",
            "incentive": m.incentive,
            "rate_card": m.rate_card,
            "band": m.band,
            "employee_id": rec_emp_id,
            "employee_name": rec_emp_name
        })
    return placements
# ── FOURTH TAB: INVOICING & BILLING ────────────────────────────

def list_invoices(db: Session) -> List[dict]:
    """
    List and synchronize invoicing records for joined candidate placements.
    """
    mappings = db.query(JobCandidateMapping).filter(
        JobCandidateMapping.status == "Joined"
    ).all()
    
    invoices_list = []
    for m in mappings:
        candidate = m.candidate
        job = m.job
        org = None
        if job.organization_id:
            org = db.query(Organization).filter(Organization.id == job.organization_id).first()
            
        invoice = db.query(Invoice).filter(Invoice.job_candidate_mapping_id == m.id).first()
        
        offered_ctc_val = 0.0
        if m.salary_offered:
            try:
                clean_ctc = "".join(c for c in m.salary_offered if c.isdigit() or c == ".")
                if clean_ctc:
                    offered_ctc_val = float(clean_ctc)
            except ValueError:
                pass
                
        if not invoice:
            # Create a default database entry for the invoice
            org_id = org.organization_id if org and org.organization_id else "UNKNOWN"
            today = date.today()
            day_str = f"{today.day:02d}"
            month_str = f"{today.month:02d}"
            year_str = f"{today.year}"
            inv_num = f"INV-{org_id}-{day_str}-{month_str}-{year_str}"
            
            invoice = Invoice(
                job_candidate_mapping_id=m.id,
                invoice_number=inv_num,
                invoice_date=today,
                billable_ctc=offered_ctc_val,
                gross=0.0,
                cgst=org.cgst if org else 0.0,
                sgst=org.sgst if org else 0.0,
                igst=org.igst if org else 0.0,
                total_gst=0.0,
                billable_amount=0.0,
                tds_deduction=0.0,
                deduction=0.0,
                received_amount=0.0,
                balance_amount=0.0,
                received_date=None,
                candidate_status="Joined",
                billing_status="Pending"
            )
            db.add(invoice)
            db.commit()
            db.refresh(invoice)
            
        # Compile response dictionary with candidate/organization details
        org_id = org.organization_id if org and org.organization_id else "UNKNOWN"
        inv_date = invoice.invoice_date or m.joining_date or date.today()
        day_str = f"{inv_date.day:02d}"
        month_str = f"{inv_date.month:02d}"
        year_str = f"{inv_date.year}"
        display_inv_num = f"INV-{org_id}-{day_str}-{month_str}-{year_str}"

        # Synchronize invoice number in DB if it was created in the old format
        if invoice.invoice_number != display_inv_num:
            invoice.invoice_number = display_inv_num
            db.commit()
            db.refresh(invoice)

        invoices_list.append({
            "id": invoice.id,
            "job_candidate_mapping_id": invoice.job_candidate_mapping_id,
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.invoice_date,
            "billable_ctc": invoice.billable_ctc,
            "gross": invoice.gross,
            "cgst": invoice.cgst,
            "sgst": invoice.sgst,
            "igst": invoice.igst,
            "total_gst": invoice.total_gst,
            "billable_amount": invoice.billable_amount,
            "tds_deduction": invoice.tds_deduction,
            "deduction": invoice.deduction,
            "received_amount": invoice.received_amount,
            "balance_amount": invoice.balance_amount,
            "received_date": invoice.received_date,
            "candidate_status": invoice.candidate_status,
            "billing_status": invoice.billing_status,
            
            # Read-only fields fetched from candidate placement and org
            "candidate_joined_date": m.joining_date,
            "candidate_name": f"{candidate.first_name} {candidate.last_name}".strip(),
            "job_designation": job.requirements[0].job_title if job.requirements else "—",
            "organization_name": org.organization_name if org else job.company_name,
            "location": org.location if org else None,
            "offered_ctc": offered_ctc_val,
            "gst_number": org.gst_number if org else None,
            "organization_id": org.organization_id if org else None
        })
        
    return invoices_list

def update_invoice(db: Session, mapping_id: int, payload: InvoiceUpdate) -> Invoice:
    """
    Update details for a candidate placement's invoice.
    """
    invoice = db.query(Invoice).filter(Invoice.job_candidate_mapping_id == mapping_id).first()
    if not invoice:
        # Create dynamically by running list_invoices to initialize all defaults
        list_invoices(db)
        invoice = db.query(Invoice).filter(Invoice.job_candidate_mapping_id == mapping_id).first()
        
    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(invoice, key, value)
        
    # Run dynamic calculations
    gross_val = invoice.gross or 0.0
    cgst_pct = invoice.cgst or 0.0
    sgst_pct = invoice.sgst or 0.0
    igst_pct = invoice.igst or 0.0
    
    cgst_amt = gross_val * (cgst_pct / 100.0)
    sgst_amt = gross_val * (sgst_pct / 100.0)
    igst_amt = gross_val * (igst_pct / 100.0)
    
    invoice.total_gst = cgst_amt + sgst_amt + igst_amt
    invoice.billable_amount = gross_val + invoice.total_gst
    
    # TDS Deduction = IGST * 10%
    invoice.tds_deduction = float(igst_amt * 0.1)
    
    tds = invoice.tds_deduction
    ded = invoice.deduction or 0.0
    rec = invoice.received_amount or 0.0
    
    invoice.balance_amount = invoice.billable_amount - tds - ded - rec
    
    db.commit()
    db.refresh(invoice)
    return invoice

def get_invoice_by_mapping_id(db: Session, mapping_id: int) -> Optional[dict]:
    """
    Get and format a single invoice by its candidate mapping ID.
    """
    invoices = list_invoices(db)
    for inv in invoices:
        if inv["job_candidate_mapping_id"] == mapping_id:
            return inv
    return None

def export_accounts_to_excel(db: Session, tab: str, search: Optional[str] = None):
    import openpyxl
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if tab == 'baseline':
        ws.title = "Salary Baseline"
        headers = [
            "Employee ID", "Employee Name", "Joining Date", 
            "CTC Offered", "Compliance", "Basic Pay", "HRA", "Status"
        ]
        ws.append(headers)
        
        # Style Header
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
        # Data
        accounts = list_accounts(db)
        if search:
            search_lower = search.lower().strip()
            accounts = [
                acc for acc in accounts 
                if search_lower in f"{acc.employee.first_name} {acc.employee.last_name}".lower()
                or search_lower in (acc.employee.employee_id or "").lower()
            ]
            
        for acc in accounts:
            ws.append([
                acc.employee.employee_id or "—",
                f"{acc.employee.first_name} {acc.employee.last_name}".strip(),
                acc.employee.date.strftime("%d/%m/%Y") if acc.employee.date else (acc.employee.date_of_joining.strftime("%d/%m/%Y") if acc.employee.date_of_joining else "—"),
                acc.employee.ctc or acc.ctc_offered or 0,
                acc.employee.compliance or "None",
                acc.basic_pay or 0,
                acc.hra or 0,
                acc.employee.status or "Inactive"
            ])
            
    elif tab == 'placements':
        ws.title = "Placements"
        headers = [
            "Approval Date", "Candidate ID", "Candidate Name", "Employee ID", 
            "Employee Name", "Org ID", "Organization Name", "Location", 
            "Job Designation", "Incentive", "Rate Card", "Band"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
        placements = list_placements(db)
        if search:
            search_lower = search.lower().strip()
            placements = [
                pl for pl in placements
                if search_lower in (pl.get("candidate_name") or "").lower()
                or search_lower in (pl.get("candidate_code") or "").lower()
                or search_lower in (pl.get("organization_name") or "").lower()
            ]
            
        for pl in placements:
            app_date = pl.get("approval_date")
            formatted_date = "—"
            if app_date:
                try:
                    formatted_date = app_date.strftime("%d/%m/%Y")
                except AttributeError:
                    formatted_date = str(app_date)
            ws.append([
                formatted_date,
                pl.get("candidate_code") or "—",
                pl.get("candidate_name") or "—",
                pl.get("employee_id") or "—",
                pl.get("employee_name") or "—",
                pl.get("organization_id") or "—",
                pl.get("organization_name") or "—",
                pl.get("location") or "—",
                pl.get("job_designation") or "—",
                pl.get("incentive") or "—",
                pl.get("rate_card") or "—",
                pl.get("band") or "—"
            ])
            
    elif tab == 'payroll':
        ws.title = "Payroll Calculations"
        headers = [
            "Joining Date", "Employee ID", "Employee Name", "Basic Pay", "HRA", 
            "Unpaid Leaves", "Unpaid Leave Amount", "PF Amount", "TDS Amount", 
            "Candidate Incentives", "Additional Incentive", "Client Incentive", 
            "Incentive Deducted", "Loan Amount", "Loan Deducted", "Net Salary Payout"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
        accounts = list_accounts(db)
        if search:
            search_lower = search.lower().strip()
            accounts = [
                acc for acc in accounts 
                if search_lower in f"{acc.employee.first_name} {acc.employee.last_name}".lower()
                or search_lower in (acc.employee.employee_id or "").lower()
            ]
            
        for acc in accounts:
            net_payout = (
                (acc.net_payable_salary or 0) + 
                (acc.hra or 0) - 
                (acc.pf or 0.0) - 
                (acc.tds or 0.0) + 
                (acc.additional_incentive or 0) + 
                (acc.client_incentive or 0) - 
                (acc.incentive_deducted or 0) + 
                (getattr(acc, 'candidate_incentives', 0) or 0) - 
                (acc.loan_deducted or 0)
            )
            ws.append([
                acc.employee.date.strftime("%d/%m/%Y") if acc.employee.date else (acc.employee.date_of_joining.strftime("%d/%m/%Y") if acc.employee.date_of_joining else "—"),
                acc.employee.employee_id or "—",
                f"{acc.employee.first_name} {acc.employee.last_name}".strip(),
                acc.basic_pay or 0,
                acc.hra or 0,
                acc.unpaid_leaves or 0,
                acc.unpaid_leave_amount or 0,
                acc.pf or 0.0,
                acc.tds or 0.0,
                getattr(acc, 'candidate_incentives', 0) or 0,
                acc.additional_incentive or 0,
                acc.client_incentive or 0,
                acc.incentive_deducted or 0,
                acc.loan_amount or 0,
                acc.loan_deducted or 0,
                net_payout
            ])
            
    elif tab == 'invoices':
        ws.title = "Organization Billing"
        headers = [
            "Month", "Candidate Joined Date", "Candidate Name", "Job Designation", 
            "Organization Name", "Location", "Offered CTC", "Billable CTC", 
            "Invoice Number", "Invoice Date", "GST Number", "Gross Amount", 
            "CGST", "SGST", "IGST", "Total GST", "Billable Amount", "TDS Deduction", 
            "Deduction", "Received Amount", "Balance Amount", "Received Date", 
            "Candidate Status", "Billing Status"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
        invoices = list_invoices(db)
        if search:
            search_lower = search.lower().strip()
            invoices = [
                inv for inv in invoices
                if search_lower in (inv.get("candidate_name") or "").lower()
                or search_lower in (inv.get("invoice_number") or "").lower()
                or search_lower in (inv.get("organization_name") or "").lower()
            ]
            
        for inv in invoices:
            joined_date = inv.get("candidate_joined_date")
            formatted_month = "—"
            formatted_joined = "—"
            if joined_date:
                try:
                    formatted_month = joined_date.strftime("%b %Y")
                    formatted_joined = joined_date.strftime("%d/%m/%Y")
                except AttributeError:
                    formatted_month = str(joined_date)
                    formatted_joined = str(joined_date)
            
            invoice_date = inv.get("invoice_date")
            formatted_invoice = "—"
            if invoice_date:
                try:
                    formatted_invoice = invoice_date.strftime("%d/%m/%Y")
                except AttributeError:
                    formatted_invoice = str(invoice_date)
                    
            received_date = inv.get("received_date")
            formatted_received = "—"
            if received_date:
                try:
                    formatted_received = received_date.strftime("%d/%m/%Y")
                except AttributeError:
                    formatted_received = str(received_date)

            gross_val = inv.get("gross") or 0.0
            cgst_val = gross_val * ((inv.get("cgst") or 0.0) / 100.0)
            sgst_val = gross_val * ((inv.get("sgst") or 0.0) / 100.0)
            igst_val = gross_val * ((inv.get("igst") or 0.0) / 100.0)

            ws.append([
                formatted_month,
                formatted_joined,
                inv.get("candidate_name") or "—",
                inv.get("job_designation") or "—",
                inv.get("organization_name") or "—",
                inv.get("location") or "—",
                inv.get("offered_ctc") or 0,
                inv.get("billable_ctc") or 0,
                inv.get("invoice_number") or "—",
                formatted_invoice,
                inv.get("gst_number") or "—",
                gross_val,
                cgst_val,
                sgst_val,
                igst_val,
                inv.get("total_gst") or 0.0,
                inv.get("billable_amount") or 0.0,
                inv.get("tds_deduction") or 0.0,
                inv.get("deduction") or 0.0,
                inv.get("received_amount") or 0.0,
                inv.get("balance_amount") or 0.0,
                formatted_received,
                inv.get("candidate_status") or "Candidate served",
                inv.get("billing_status") or "Pending"
            ])
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def close_salary_cycle(db: Session):
    """
    Snapshots the current payroll calculations into PayrollHistory
    and resets active monthly variables for the next cycle.
    """
    from app.models.account import PayrollCalculationsHistory, OrganizationBillingHistory, CandidatesHiredForOrganizationsHistory
    
    # Query all active accounts (list_accounts ensures they are computed & sync'd)
    accounts = list_accounts(db)
    
    # Common month year identifier for the closed cycle
    _, default_end_date = get_current_salary_cycle_range(date.today(), date(2020, 1, 1))
    common_cycle_month_year = default_end_date.strftime("%B %Y")
    
    # Snapshot active placements to CandidatesHiredForOrganizationsHistory
    db.query(CandidatesHiredForOrganizationsHistory).filter(CandidatesHiredForOrganizationsHistory.cycle_month_year == common_cycle_month_year).delete()
    placements = list_placements(db)
    for pl in placements:
        hist_pl = CandidatesHiredForOrganizationsHistory(
            approval_date=pl.get("approval_date"),
            candidate_code=pl.get("candidate_code"),
            candidate_name=pl.get("candidate_name") or "—",
            organization_id=pl.get("organization_id"),
            organization_name=pl.get("organization_name") or "—",
            location=pl.get("location"),
            job_designation=pl.get("job_designation") or "—",
            incentive=pl.get("incentive") or 0.0,
            rate_card=pl.get("rate_card"),
            band=pl.get("band"),
            employee_id=pl.get("employee_id"),
            employee_name=pl.get("employee_name"),
            cycle_month_year=common_cycle_month_year
        )
        db.add(hist_pl)

    # Snapshot active invoices to OrganizationBillingHistory
    db.query(OrganizationBillingHistory).filter(OrganizationBillingHistory.cycle_month_year == common_cycle_month_year).delete()
    invoices = list_invoices(db)
    for inv in invoices:
        gross_val = inv.get("gross") or 0.0
        cgst_val = inv.get("cgst") or 0.0
        sgst_val = inv.get("sgst") or 0.0
        igst_val = inv.get("igst") or 0.0
        cgst_amt = gross_val * (cgst_val / 100.0)
        sgst_amt = gross_val * (sgst_val / 100.0)
        igst_amt = gross_val * (igst_val / 100.0)
        total_gst = cgst_amt + sgst_amt + igst_amt
        billable_amount = gross_val + total_gst
        tds_deduction = igst_amt * 0.1
        
        hist_inv = OrganizationBillingHistory(
            job_candidate_mapping_id=inv.get("id"),
            candidate_name=inv.get("candidate_name") or "—",
            job_designation=inv.get("job_designation") or "—",
            organization_name=inv.get("organization_name") or "—",
            location=inv.get("location"),
            offered_ctc=inv.get("offered_ctc") or 0.0,
            billable_ctc=inv.get("billable_ctc") or 0.0,
            invoice_number=inv.get("invoice_number") or "—",
            invoice_date=inv.get("invoice_date"),
            gst_number=inv.get("gst_number"),
            gross=gross_val,
            cgst=cgst_val,
            sgst=sgst_val,
            igst=igst_val,
            total_gst=total_gst,
            billable_amount=billable_amount,
            tds_deduction=tds_deduction,
            deduction=inv.get("deduction") or 0.0,
            received_amount=inv.get("received_amount") or 0.0,
            balance_amount=billable_amount - tds_deduction - (inv.get("deduction") or 0.0) - (inv.get("received_amount") or 0.0),
            received_date=inv.get("received_date"),
            candidate_status=inv.get("candidate_status"),
            billing_status=inv.get("billing_status"),
            cycle_month_year=common_cycle_month_year
        )
        db.add(hist_inv)
    
    for acc in accounts:
        joining_date = acc.employee.date or acc.employee.date_of_joining
        start_date, end_date = get_current_salary_cycle_range(date.today(), joining_date)
        cycle_month_year = end_date.strftime("%B %Y")  # e.g., "July 2026"
        
        # Calculate net payout: Calculated Basic Pay + HRA - PF - TDS + Additional Incentive + Client Incentive - Incentive Deducted + Candidate Incentives - Loan Deducted
        net_payout = (
            (acc.net_payable_salary or 0) + 
            (acc.hra or 0) - 
            (acc.pf or 0.0) - 
            (acc.tds or 0.0) + 
            (acc.additional_incentive or 0) + 
            (acc.client_incentive or 0) - 
            (acc.incentive_deducted or 0) + 
            (getattr(acc, 'candidate_incentives', 0) or 0) - 
            (acc.loan_deducted or 0)
        )
        
        # Remove any existing snapshot for this employee/month to prevent duplicates
        db.query(PayrollCalculationsHistory).filter(
            PayrollCalculationsHistory.employee_id == acc.employee_id,
            PayrollCalculationsHistory.cycle_month_year == cycle_month_year
        ).delete()
        
        # Insert snapshot
        history_entry = PayrollCalculationsHistory(
            employee_id=acc.employee_id,
            employee_name=f"{acc.employee.first_name} {acc.employee.last_name}".strip(),
            cycle_start_date=start_date,
            cycle_end_date=end_date,
            cycle_month_year=cycle_month_year,
            basic_pay=acc.basic_pay or 0,
            hra=acc.hra or 0,
            unpaid_leaves=acc.unpaid_leaves or 0.0,
            unpaid_leave_amount=acc.unpaid_leave_amount or 0,
            pf=acc.pf or 0.0,
            tds=acc.tds or 0.0,
            incentives=acc.incentives or 0,
            candidate_incentives=getattr(acc, 'candidate_incentives', 0) or 0,
            additional_incentive=acc.additional_incentive or 0,
            client_incentive=acc.client_incentive or 0,
            incentive_deducted=acc.incentive_deducted or 0,
            loan_amount=acc.loan_amount or 0,
            loan_deducted=acc.loan_deducted or 0,
            deduction_amount=acc.deduction_amount or 0,
            net_salary_pay=net_payout
        )
        db.add(history_entry)
        
        # Reset monthly input fields for the active account
        acc.additional_incentive = 0
        acc.client_incentive = 0
        acc.incentive_deducted = 0
        acc.loan_deducted = 0
        acc.deduction_amount = 0
        
    db.commit()
    return {"success": True, "message": "Month closed successfully, calculations reset."}

def list_history_months(db: Session) -> List[str]:
    """
    List all unique months saved in payroll history.
    """
    from app.models.account import PayrollCalculationsHistory
    results = db.query(PayrollCalculationsHistory.cycle_month_year).distinct().all()
    # Flat list
    months = [r[0] for r in results if r[0]]
    # Sort months chronologically by parsing them
    from datetime import datetime
    try:
        months.sort(key=lambda m: datetime.strptime(m, "%B %Y"), reverse=True)
    except ValueError:
        months.sort(reverse=True)
    return months

def export_history_to_excel(db: Session, month_year: str):
    """
    Export payroll history snapshots for a given month as an Excel file.
    """
    import openpyxl
    from io import BytesIO
    from app.models.account import PayrollCalculationsHistory
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {month_year}"
    
    headers = [
        "Joining Date", "Employee ID", "Employee Name", "Basic Pay", "HRA", 
        "Unpaid Leaves", "Unpaid Leave Amount", "PF Amount", "TDS Amount", 
        "Candidate Incentives", "Additional Incentive", "Client Incentive", 
        "Incentive Deducted", "Loan Amount", "Loan Deducted", "Net Salary Payout"
    ]
    ws.append(headers)
    
    # Style Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
    history_records = db.query(PayrollCalculationsHistory).filter(
        PayrollCalculationsHistory.cycle_month_year == month_year
    ).all()
    
    for r in history_records:
        ws.append([
            r.cycle_start_date.strftime("%d/%m/%Y"),
            r.employee.employee_id if r.employee else "—",
            r.employee_name,
            r.basic_pay,
            r.hra,
            r.unpaid_leaves,
            r.unpaid_leave_amount,
            r.pf,
            r.tds,
            getattr(r, 'candidate_incentives', 0) or 0,
            r.additional_incentive,
            r.client_incentive,
            r.incentive_deducted,
            r.loan_amount,
            r.loan_deducted,
            r.net_salary_pay
        ])
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_billing_history_to_excel(db: Session, month_year: str):
    """
    Export organization billing history snapshots for a given month as an Excel file.
    """
    import openpyxl
    from io import BytesIO
    from app.models.account import OrganizationBillingHistory
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Billing History"
    
    headers = [
        "Month", "Candidate Joined Date", "Candidate Name", "Job Designation", 
        "Organization Name", "Location", "Offered CTC", "Billable CTC", 
        "Invoice Number", "Invoice Date", "GST Number", "Gross", 
        "CGST Amount", "SGST Amount", "IGST Amount", "Total GST", 
        "Billable Amount", "TDS Deduction", "Deduction", "Received Amount", 
        "Balance Amount", "Received Date", "Candidate Status", "Billing Status"
    ]
    ws.append(headers)
    
    # Style Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
    records = db.query(OrganizationBillingHistory).filter(
        OrganizationBillingHistory.cycle_month_year == month_year
    ).all()
    
    for r in records:
        formatted_joined = r.invoice_date.strftime("%d/%m/%Y") if r.invoice_date else "—"
        formatted_invoice = r.invoice_date.strftime("%d/%m/%Y") if r.invoice_date else "—"
        formatted_received = r.received_date.strftime("%d/%m/%Y") if r.received_date else "—"
        
        gross_val = r.gross or 0.0
        cgst_amt = gross_val * ((r.cgst or 0.0) / 100.0)
        sgst_amt = gross_val * ((r.sgst or 0.0) / 100.0)
        igst_amt = gross_val * ((r.igst or 0.0) / 100.0)
        
        ws.append([
            r.cycle_month_year,
            formatted_joined,
            r.candidate_name,
            r.job_designation,
            r.organization_name,
            r.location or "—",
            r.offered_ctc,
            r.billable_ctc,
            r.invoice_number,
            formatted_invoice,
            r.gst_number or "—",
            gross_val,
            cgst_amt,
            sgst_amt,
            igst_amt,
            r.total_gst,
            r.billable_amount,
            r.tds_deduction,
            r.deduction,
            r.received_amount,
            r.balance_amount,
            formatted_received,
            r.candidate_status or "Candidate served",
            r.billing_status or "Pending"
        ])
        
    # Auto-fit columns
    for col in ws.columns:
      max_len = max(len(str(cell.value or '')) for cell in col)
      col_letter = openpyxl.utils.get_column_letter(col[0].column)
      ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_bank_credit_details(db: Session, bank_name: str):
    """
    Export bank credit details of employees to Excel depending on the chosen bank:
    - HDFC: employees where compliance == None (or empty) using 28 HDFC columns format
    - ICICI: employees where compliance == PF or compliance == TDS
    """
    import openpyxl
    from io import BytesIO
    
    # Sync and fetch all accounts (this keeps calculations up to date in DB)
    active_accounts = list_accounts(db)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{bank_name} Credit Details"
    
    if bank_name.upper() == "HDFC":
        headers = [
            'Transaction Type', 'Blank', 'Beneficiary Account Number', 'Instrument Amount', 
            'Beneficiary Name', 'Blank', 'Blank', 'Blank', 'Blank', 'Blank', 
            'Blank', 'Blank', 'Instruction Reference Number', 'Customer Reference Number', 
            'Payment details 1', 'Blank', 'Blank', 'Blank', 'Blank', 'Blank', 
            'Blank', 'Blank', 'DATE - DD/MM/YYYY', 'Blank', 'IFC Code', 'Blank', 
            'Blank', 'Beneficiary email id'
        ]
    else:
        # ICICI
        headers = [
            "Debit Ac No", "Beneficiary Ac No", "Beneficiary Name", "Amt", "Pay Mod", 
            "Date", "IFSC", "Payable Location", "Print Location", "Bene Mobile No.", 
            "Bene Email ID", "Bene add1", "Bene add2", "Bene add3", "Bene add4", 
            "Add Details 1", "Add Details 2", "Add Details 3", "Add Details 4", 
            "Add Details 5", "Reamrks"
        ]
        
    ws.append(headers)
    
    # Style Header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
    for acc in active_accounts:
        emp = acc.employee
        if not emp:
            continue
            
        comp = (emp.compliance or "").strip().upper()
        match = False
        if bank_name.upper() == "HDFC":
            if comp in ("", "NONE", "NULL"):
                match = True
        elif bank_name.upper() == "ICICI":
            if comp in ("PF", "TDS"):
                match = True
                
        if match:
            fullname = f"{emp.first_name or ''} {emp.last_name or ''}".strip() or "—"
            today_str = date.today().strftime("%d/%m/%Y")
            if bank_name.upper() == "HDFC":
                ws.append([
                    "NEFT",                               # Transaction Type
                    "",                                   # Blank
                    emp.bank_account_number or "",        # Beneficiary Account Number
                    acc.net_salary_pay or 0,              # Instrument Amount
                    fullname,                             # Beneficiary Name
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    f"REF-{emp.employee_id or emp.id}",   # Instruction Reference Number
                    f"CUST-{emp.employee_id or emp.id}",  # Customer Reference Number
                    "Salary Payout",                      # Payment details 1
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    "",                                   # Blank
                    today_str,                            # DATE - DD/MM/YYYY
                    "",                                   # Blank
                    emp.bank_ifsc_code or "",             # IFC Code
                    "",                                   # Blank
                    "",                                   # Blank
                    emp.email or ""                       # Beneficiary email id
                ])
            else:
                ws.append([
                    "",                                   # Debit Ac No
                    emp.bank_account_number or "",        # Beneficiary Ac No
                    fullname,                             # Beneficiary Name
                    acc.net_salary_pay or 0,              # Amt
                    "NEFT",                               # Pay Mod
                    today_str,                            # Date
                    emp.bank_ifsc_code or "",             # IFSC
                    "",                                   # Payable Location
                    "",                                   # Print Location
                    emp.contact_number or "",             # Bene Mobile No.
                    emp.email or "",                      # Bene Email ID
                    "",                                   # Bene add1
                    "",                                   # Bene add2
                    "",                                   # Bene add3
                    "",                                   # Bene add4
                    "Salary Payout",                      # Add Details 1
                    "",                                   # Add Details 2
                    "",                                   # Add Details 3
                    "",                                   # Add Details 4
                    "",                                   # Add Details 5
                    "Salary"                              # Reamrks
                ])
                
    # Auto-fit columns
    for col in ws.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
