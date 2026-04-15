import sys
import os
import openpyxl
from sqlalchemy.orm import Session

# Add backend directory to path to import app modules
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.core.database import SessionLocal
from app.models.organization import Organization
from app.services.organization_service import generate_organization_id, check_organization_exists
from app.schemas.organization import OrganizationCreate

def import_from_excel(file_path: str):
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        return

    db: Session = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Assume first column is organization_name
        # Skip header if it exists. We check if first row looks like a header.
        start_row = 1
        first_val = sheet.cell(row=1, column=1).value
        if first_val and (str(first_val).lower() == 'organization_name' or str(first_val).lower() == 'name'):
            start_row = 2
            print(f"Detected header: {first_val}")

        imported_count = 0
        skipped_count = 0

        for row in range(start_row, sheet.max_row + 1):
            org_name = sheet.cell(row=row, column=1).value
            if not org_name:
                continue
            
            org_name = str(org_name).strip()
            if not org_name:
                continue

            # Check for duplicates in DB (case-insensitive)
            if check_organization_exists(db, org_name):
                print(f"Skipping duplicate: '{org_name}'")
                skipped_count += 1
                continue

            # Check for duplicates in the current session (to avoid multi-row duplicates in same excel)
            # though check_organization_exists uses db.query, it won't see uncommitted ones yet.
            # But the service actually does commit inside, wait, let's check organization_service.py.
            # create_organization(db, payload) commits.

            try:
                # Use generate_organization_id
                org_id = generate_organization_id(db)
                
                # The schema requires commission_percentage. Setting default 0.0
                new_org = Organization(
                    organization_id=org_id,
                    organization_name=org_name,
                    status='in_progress',
                    commission_percentage=0.0
                )
                db.add(new_org)
                db.commit() # Commit each time to ensure generate_organization_id gets next ID correctly
                db.refresh(new_org)
                
                print(f"Imported: '{org_name}' -> {org_id}")
                imported_count += 1
            except Exception as e:
                db.rollback()
                print(f"Error importing '{org_name}': {e}")
                skipped_count += 1

        print(f"\nImport Summary:")
        print(f"Total Imported: {imported_count}")
        print(f"Total Skipped: {skipped_count}")

    finally:
        db.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_organizations.py <path_to_excel_file>")
        # Create a sample file if none provided for demo
        file_path = "companies.xlsx"
        if not os.path.exists(file_path):
            print("Creating sample companies.xlsx for demonstration...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'organization_name'
            ws['A2'] = 'TCS'
            ws['A3'] = 'Infosys'
            ws['A4'] = 'Wipro'
            ws['A5'] = 'Accenture'
            ws['A6'] = '  TCS  ' # Duplicate check test
            ws['A7'] = '' # Empty row test
            wb.save(file_path)
            print(f"Sample file created: {file_path}")
            import_from_excel(file_path)
        else:
            print(f"Using existing {file_path}")
            import_from_excel(file_path)
    else:
        import_from_excel(sys.argv[1])
